# # Reading an excel file using Python
import openpyxl
from openpyxl.styles import Alignment, Font

# Class Trainee that stores necessary information about the trainee, including
# name of the trainee and their host (tname and hname), the date and type of
# the submitted evaluation (date and type)
class Trainee:
    # constructor
    def __init__(self, type, tname, hname, date):
        self.type = type
        self.tname= tname
        self.hname = hname
        self.date = date
        self.update = 0
    # attribute update becomes 1 when the evaluation is updated in the sheet
    def done_update(self):
        self.update = 1

# Columns corresponding to each type of evaluation
# remainder: type_of_email = ["Host Trainer Mid-Point", "Host Trainer Final", "Trainee/Intern Initial",
#                                 "Trainee/Intern Mid-Point", "Trainee/Intern Final"]
type_col = ['O','P','L','M','N']

# Update the sheet at position cell with the date as the info
# font with size 10 and at the center of the cell
def write_format(ws, cell, info):
    ws[cell] = info
    ws[cell].font = Font(size=10)
    ws[cell].alignment = Alignment(horizontal='center',vertical='center')


# function add_dates that goes through a list of trainees, search for them
# in the XLSX sheet at loc location. If the trainee is found, update the date
# attribute of the trainee in the correct column according to their evalution
# type. If found, the update attribute of that trainee needs to be updated to 1
def add_dates(trainees, loc):
    index = 1;
    # To open Workbook
    wb = openpyxl.load_workbook(loc)
    for trainee in trainees:
        # Look through two sheets
        # The first sheet includes trainees still in the US
        # The second sheet includes trainees left the US already
        for ws in wb:
            # Method 2: 20/20 success on training
            # A dict to store the column that each key word be found in
            # Key is the portion of the trainee name or the host name
            # Value is the list of column that the key is found in
            # Format: solution = {"key1:[column1,column2...]","key2:[column1,column2...]",...}
            solutions = {}

            # Add key and value to solution list
            # First with the trainee name list
            # Note that the trainee name is in column A of the sheet (column=1)
            for k in trainee.tname:
                for j in range(2,ws.max_row+1):
                    if str(ws.cell(row=j, column=1).value).find(k) != -1:
                        if k not in solutions.keys():
                            solutions[k] = [j]
                        else:
                            solutions[k].append(j)
            # Second with the host name list
            # Note that the host name is in column D of the sheet (column=4)
            for k in trainee.hname:
                for j in range(2,ws.max_row+1):
                    if str(ws.cell(row=j, column=4).value).find(k) != -1:
                        if k not in solutions.keys():
                            solutions[k] = [j]
                        else:
                            solutions[k].append(j)
            # Now make a new dict named possibles with the column as the key
            # and value is the list of name portion that can be found in the column
            possibles = {};
            for key in solutions.keys():
                for value in solutions[key]:
                    if value not in possibles:
                        possibles[value] = [key]
                    else:
                        possibles[value].append(key)

            # Evaluate to see which column matches the most name portion
            # max_pos is the column with the most matches
            # max_len helps find the column with the most matches
            max_pos = 0
            max_len = 0

            for key in possibles.keys():
                if len(possibles[key]) > max_len:
                    check = True;
                    # Check for extra requirements
                    # Trainee name at the cell looking at
                    tname_in_cell = str(ws.cell(row=key,column=1).value)
                    # Host name at the cell looking at
                    hname_in_cell = str(ws.cell(row=key,column=4).value)
                    # Check if the first two letter of each name matches with
                    # the name in the sheet
                    for index2 in range(0,len(trainee.tname)):
                        name = trainee.tname[index2]
                        if len(name)>=3:
                            if tname_in_cell.find(trainee.tname[len(trainee.tname)-1][:2])==-1:
                                check = False
                                break
                # If all requirements satified, update max_pos and max_len and break
                if check == True:
                    max_pos = key
                    max_len = len(possibles[key])
                    break
            # if found, do not need to consider the other sheet, thus break
            if max_pos!=0:
                break
        # If there is a max_pos (max_pos!=0), the update the trainee and add
        # the date into the worksheet using the write_format function
        # Else, print can't find
        if max_pos!=0:
            # Update that the trainee is updated
            trainee.done_update()
            cell = type_col[trainee.type]+str(max_pos)
            write_format(ws,cell,trainee.date)
        else:
            print(index, ". Can't find")
        index+=1


            # # Method 1: Success 12/20
            # # Position list of all the cell in column A that has trainee fist name
            # pos_find = []
            # pos_host = []
            # # Iterate all the rows in column 1 to find which rows have the trainee first name and last name
            # for j in range(2,ws.max_row):
            #     # print(str(ws.cell(row=j, column=1).value))
            #     if str(ws.cell(row=j, column=1).value).find(trainee.tname[0]) != -1:
            #         if str(ws.cell(row=j, column=1).value).find(trainee.tname[len(trainee.tname)-1]) != -1:
            #             pos_find.append(j)  # add the position to the list if first name found
            # # if there is only one position found
            # if len(pos_find) == 1:
            #     # Update that the trainee is updated
            #     trainee.done_update()
            #     cell = type_col[trainee.type]+str(pos_find[0])
            #     write_format(ws,cell,trainee.date)
            #     break
            # # if there is more than one position, check the host name
            # elif len(pos_find) >1:
            #     for j in pos_find():
            #         if str(ws.cell(row=j, column=4).value).find(trainee.hname[0]) != -1:
            #             pos_host.append(j)
            #     if len(pos_host) == 1:
            #         # Update that the trainee is updated
            #         trainee.done_update()
            #         cell = type_col[trainee.type]+str(pos_find[0])
            #         write_format(ws,cell,trainee.date)
            #         break
            #
            # if len(pos_find) == 1 or len(pos_host) ==1:
            #     break

    wb.save("new.xlsx")
    return 0
